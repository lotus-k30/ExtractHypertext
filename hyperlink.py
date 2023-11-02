""" 
Author:     Dr. Kamaljeet Kaur
Date:       02-11-2023
Status:     Tested OK
Description: This code edits the excel sheet
            The hyperlink contained behind the text in column D of Sheet 1
            is extracted and pasted into column E of the workbook.
Input:      Excel file 'school_data.xlsx'
Output:     Modified Excel file 'school_data.xlsx'
"""

import openpyxl
# Open workbook
wb = openpyxl.load_workbook('school_data.xlsx')
# Select the desired worksheet
ws = wb['Sheet1']
# Loop to copy cells one by one
for r in range(1,ws.max_row):
    # Extract hyperlink from column D and paste in Column E
    hyperlink = ws["D"+str(r)].hyperlink
    ws['E'+str(r)].value = str(hyperlink.display)
# Save the workbook
wb.save('school_data.xlsx')
# Display msg to show that task has been accomplished
print("Writing Done")
